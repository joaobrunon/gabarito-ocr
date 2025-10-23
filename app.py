#!/usr/bin/env python3
"""
Aplicação Web para Correção de Gabaritos
Sistema completo de upload, correção e visualização de relatórios
"""

from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import json
import os
from pathlib import Path
from datetime import datetime
import subprocess
import sys
from gerador_gabarito import GeradorGabarito
from gerar_gabaritos_personalizados import ler_csv_alunos, listar_turmas, gerar_gabaritos_turma
import csv
import zipfile
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import unquote

# Configuração
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max
app.config['UPLOAD_FOLDER'] = 'pdfs_para_corrigir'
app.config['REPORTS_FOLDER'] = 'relatorios_correcao'
app.config['GABARITOS_FOLDER'] = 'gabaritos'
app.config['GABARITOS_PDF_FOLDER'] = 'gabaritos_gerados'
app.config['CSV_UPLOAD_FOLDER'] = 'uploads_csv'
app.config['CSV_ALUNOS_FOLDER'] = 'csv_alunos_referencia'

# Criar pastas se não existirem
Path(app.config['UPLOAD_FOLDER']).mkdir(exist_ok=True)
Path(app.config['REPORTS_FOLDER']).mkdir(exist_ok=True)
Path(app.config['GABARITOS_FOLDER']).mkdir(exist_ok=True)
Path(app.config['GABARITOS_PDF_FOLDER']).mkdir(exist_ok=True)
Path(app.config['CSV_UPLOAD_FOLDER']).mkdir(exist_ok=True)
Path(app.config['CSV_ALUNOS_FOLDER']).mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf'}
ALLOWED_CSV_EXTENSIONS = {'csv'}
GABARITO_OFICIAL = 'gabarito_oficial.json'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_csv_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_CSV_EXTENSIONS

@app.route('/')
def index():
    """Página principal"""
    return render_template('index.html')

@app.route('/api/reports', methods=['GET'])
def get_reports():
    """Retorna lista de relatórios"""
    reports = []
    reports_path = Path(app.config['REPORTS_FOLDER'])

    # Verificar se há filtro de turma
    filtro_turma = request.args.get('turma', '').strip()

    for json_file in sorted(reports_path.glob('*_relatorio.json'), reverse=True):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            turma = data.get('identificacao', {}).get('turma', 'Sem turma')

            # Aplicar filtro se especificado
            if filtro_turma and turma != filtro_turma:
                continue

            reports.append({
                'nome': data.get('identificacao', {}).get('nome', 'Desconhecido'),
                'turma': turma,
                'data': data.get('data_correcao', 'N/A'),
                'nota': data.get('nota', 0),
                'acertos': data.get('acertos', 0),
                'erros': data.get('erros', 0),
                'total': data.get('total_questoes', 40),
                'json_file': json_file.name,
                'html_file': json_file.with_suffix('.html').name
            })
        except:
            pass

    return jsonify(reports)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Recebe PDF e começa a corrigir"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Arquivo vazio'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Apenas arquivos PDF são permitidos'}), 400

    # Obter gabarito selecionado
    gabarito_nome = request.form.get('gabarito', '').strip()
    if not gabarito_nome:
        return jsonify({'error': 'Selecione um gabarito'}), 400

    # Verificar se o gabarito existe
    gabarito_path = Path(app.config['GABARITOS_FOLDER']) / f"{gabarito_nome}.json"
    if not gabarito_path.exists():
        return jsonify({'error': f'Gabarito "{gabarito_nome}" não encontrado'}), 404

    # Salvar arquivo
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    # Corrigir PDF com o gabarito especificado
    try:
        # Usar o python3 do venv atual
        base_dir = os.path.dirname(os.path.abspath(__file__))
        venv_python = os.path.join(base_dir, 'venv', 'bin', 'python3')
        corrigir_script = os.path.join(base_dir, 'corrigir_rapido.py')

        result = subprocess.run(
            [venv_python, corrigir_script, filepath, str(gabarito_path)],
            capture_output=True,
            text=True,
            timeout=30
        )

        if result.returncode != 0:
            return jsonify({'error': 'Erro ao corrigir PDF'}), 500

        # Procurar relatório gerado
        base_name = Path(filename).stem
        json_files = list(Path(app.config['REPORTS_FOLDER']).glob(f'*{base_name}*relatorio.json'))

        if not json_files:
            return jsonify({'error': 'Relatório não gerado'}), 500

        json_file = json_files[0]

        # Carregar dados do relatório
        with open(json_file, 'r', encoding='utf-8') as f:
            report_data = json.load(f)

        # Tentar identificar turma do aluno
        turma = encontrar_turma_aluno(filename)
        if turma:
            # Adicionar turma ao relatório JSON e salvar
            if 'identificacao' not in report_data:
                report_data['identificacao'] = {}
            report_data['identificacao']['turma'] = turma

            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2)

            # Regerar HTML com a turma atualizada
            try:
                from visualizar_relatorio import gerar_html_relatorio
                gerar_html_relatorio(str(json_file))
            except:
                pass

        return jsonify({
            'success': True,
            'message': 'PDF corrigido com sucesso!',
            'report': {
                'nome': report_data.get('identificacao', {}).get('nome', 'Desconhecido'),
                'turma': report_data.get('identificacao', {}).get('turma', 'N/A'),
                'data': report_data.get('data_correcao', 'N/A'),
                'nota': report_data.get('nota', 0),
                'acertos': report_data.get('acertos', 0),
                'erros': report_data.get('erros', 0),
                'em_branco': report_data.get('em_branco', 0),
                'total': report_data.get('total_questoes', 40),
                'percentual': report_data.get('percentual', 0),
                'json_file': json_file.name
            }
        })

    except subprocess.TimeoutExpired:
        return jsonify({'error': 'Timeout ao corrigir PDF'}), 500
    except Exception as e:
        return jsonify({'error': f'Erro: {str(e)}'}), 500

@app.route('/api/report/<filename>')
def get_report(filename):
    """Retorna dados de um relatório específico"""
    json_path = Path(app.config['REPORTS_FOLDER']) / filename

    if not json_path.exists():
        return jsonify({'error': 'Relatório não encontrado'}), 404

    try:
        # Se tiver parâmetro download, fazer download do arquivo
        if request.args.get('download'):
            return send_file(json_path, as_attachment=True, download_name=filename)

        # Caso contrário, retornar JSON
        with open(json_path, 'r', encoding='utf-8') as f:
            report = json.load(f)
        return jsonify(report)
    except:
        return jsonify({'error': 'Erro ao ler relatório'}), 500

@app.route('/api/stats')
def get_stats():
    """Retorna estatísticas gerais"""
    stats = {
        'total_relatorios': 0,
        'media_nota': 0,
        'maior_nota': 0,
        'menor_nota': 10,
        'por_turma': {}
    }

    reports_path = Path(app.config['REPORTS_FOLDER'])
    notas = []
    notas_por_turma = {}

    for json_file in reports_path.glob('*_relatorio.json'):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            nota = data.get('nota', 0)
            turma = data.get('identificacao', {}).get('turma', 'Sem turma')

            notas.append(nota)

            # Acumular por turma
            if turma not in notas_por_turma:
                notas_por_turma[turma] = []
            notas_por_turma[turma].append(nota)
        except:
            pass

    if notas:
        stats['total_relatorios'] = len(notas)
        stats['media_nota'] = round(sum(notas) / len(notas), 2)
        stats['maior_nota'] = max(notas)
        stats['menor_nota'] = min(notas)

    # Calcular estatísticas por turma
    for turma, notas_turma in notas_por_turma.items():
        if notas_turma:
            stats['por_turma'][turma] = {
                'total': len(notas_turma),
                'media': round(sum(notas_turma) / len(notas_turma), 2),
                'maior': max(notas_turma),
                'menor': min(notas_turma)
            }

    return jsonify(stats)

@app.route('/api/envios')
def get_envios():
    """Retorna envios agrupados por turma e data"""
    envios_por_turma = {}

    reports_path = Path(app.config['REPORTS_FOLDER'])

    for json_file in reports_path.glob('*_relatorio.json'):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # Extrair informações
            identificacao = data.get('identificacao', {})
            nome = identificacao.get('nome', 'Desconhecido')
            matricula = identificacao.get('matricula', '')
            turma = identificacao.get('turma', 'Sem turma')
            nota = data.get('nota', 0)
            acertos = data.get('acertos', 0)
            total = data.get('total_questoes', 0)

            # Data do relatório (usar timestamp do arquivo)
            timestamp = json_file.stat().st_mtime
            data_envio = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')

            # Organizar por turma
            if turma not in envios_por_turma:
                envios_por_turma[turma] = {
                    'turma': turma,
                    'total_alunos': 0,
                    'media_nota': 0,
                    'alunos': []
                }

            # Adicionar aluno
            envios_por_turma[turma]['alunos'].append({
                'nome': nome,
                'matricula': matricula,
                'nota': nota,
                'acertos': acertos,
                'total': total,
                'percentual': round((acertos / total * 100) if total > 0 else 0, 1),
                'data_envio': data_envio,
                'relatorio_json': json_file.name,
                'relatorio_html': json_file.name.replace('.json', '.html')
            })

        except Exception as e:
            print(f"Erro ao processar {json_file}: {e}")
            continue

    # Calcular estatísticas por turma e ordenar alunos
    resultado = []
    for turma, dados in envios_por_turma.items():
        # Ordenar alunos por nome
        dados['alunos'].sort(key=lambda x: x['nome'])

        # Calcular estatísticas
        dados['total_alunos'] = len(dados['alunos'])
        if dados['alunos']:
            notas = [a['nota'] for a in dados['alunos']]
            dados['media_nota'] = round(sum(notas) / len(notas), 2)
            dados['maior_nota'] = max(notas)
            dados['menor_nota'] = min(notas)

        resultado.append(dados)

    # Ordenar turmas por nome
    resultado.sort(key=lambda x: x['turma'])

    return jsonify(resultado)

@app.route('/api/envios/export/<turma>')
def export_turma_excel(turma):
    """Exporta relatório da turma para Excel"""
    try:
        # Decodificar nome da turma da URL
        turma_nome = unquote(turma)

        reports_path = Path(app.config['REPORTS_FOLDER'])
        alunos_turma = []

        # Coletar todos os alunos da turma
        for json_file in reports_path.glob('*_relatorio.json'):
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                identificacao = data.get('identificacao', {})
                turma_aluno = identificacao.get('turma', 'Sem turma')

                if turma_aluno == turma_nome:
                    # Coletar todas as informações necessárias
                    nome = identificacao.get('nome', 'Desconhecido')
                    matricula = identificacao.get('matricula', '')
                    nota = data.get('nota', 0)
                    acertos = data.get('acertos', 0)
                    erros = data.get('erros', 0)
                    total = data.get('total_questoes', 0)
                    percentual = round((acertos / total * 100) if total > 0 else 0, 1)

                    # Detalhes das questões
                    detalhes = data.get('detalhes', [])

                    alunos_turma.append({
                        'nome': nome,
                        'matricula': matricula,
                        'acertos': acertos,
                        'erros': erros,
                        'total': total,
                        'percentual': percentual,
                        'nota': nota,
                        'detalhes': detalhes
                    })

            except Exception as e:
                print(f"Erro ao processar {json_file}: {e}")
                continue

        if not alunos_turma:
            return jsonify({'error': 'Nenhum aluno encontrado para esta turma'}), 404

        # Ordenar alunos por nome
        alunos_turma.sort(key=lambda x: x['nome'])

        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = turma_nome[:31]  # Excel limita a 31 caracteres

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"Relatório de Correção - {turma_nome}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Cabeçalhos
        headers = ['Nome do Aluno', 'RA', 'Acertos', 'Erros', 'Total', 'Percentual (%)', 'Nota']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Dados dos alunos
        current_row = 4
        for aluno in alunos_turma:
            ws.cell(row=current_row, column=1, value=aluno['nome']).border = border
            ws.cell(row=current_row, column=2, value=aluno['matricula']).border = border
            ws.cell(row=current_row, column=3, value=aluno['acertos']).border = border
            ws.cell(row=current_row, column=4, value=aluno['erros']).border = border
            ws.cell(row=current_row, column=5, value=aluno['total']).border = border

            perc_cell = ws.cell(row=current_row, column=6, value=aluno['percentual'])
            perc_cell.border = border
            perc_cell.alignment = Alignment(horizontal='center')

            # Colorir percentual
            if aluno['percentual'] >= 70:
                perc_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif aluno['percentual'] >= 50:
                perc_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                perc_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            nota_cell = ws.cell(row=current_row, column=7, value=round(aluno['nota'], 1))
            nota_cell.border = border
            nota_cell.font = Font(bold=True)

            current_row += 1

        # Estatísticas
        current_row += 2
        ws.cell(row=current_row, column=1, value="ESTATÍSTICAS DA TURMA").font = Font(bold=True, size=12)
        current_row += 1

        total_alunos = len(alunos_turma)
        media_turma = round(sum(a['nota'] for a in alunos_turma) / total_alunos, 2) if total_alunos > 0 else 0

        stats = [
            ('Total de Alunos:', total_alunos),
            ('Média da Turma:', media_turma),
        ]

        for label, value in stats:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

        # Ajustar larguras
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10

        # Salvar em BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Nome do arquivo
        filename = f"Relatorio_{turma_nome.replace(' ', '_')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao gerar relatório: {str(e)}'}), 500

@app.route('/api/envios/limpar', methods=['POST'])
def limpar_envios():
    """Limpa todos os relatórios de correção"""
    try:
        reports_path = Path(app.config['REPORTS_FOLDER'])

        # Contar arquivos antes de deletar
        json_files = list(reports_path.glob('*_relatorio.json'))
        html_files = list(reports_path.glob('*_relatorio.html'))

        total_arquivos = len(json_files) + len(html_files)

        # Deletar arquivos JSON
        for json_file in json_files:
            json_file.unlink()

        # Deletar arquivos HTML
        for html_file in html_files:
            html_file.unlink()

        return jsonify({
            'success': True,
            'message': f'{total_arquivos} arquivos deletados com sucesso',
            'json_deletados': len(json_files),
            'html_deletados': len(html_files)
        })

    except Exception as e:
        print(f"Erro ao limpar envios: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao limpar envios: {str(e)}'}), 500

@app.route('/relatorio/<filename>')
def view_report(filename):
    """Retorna o HTML do relatório"""
    html_path = Path(app.config['REPORTS_FOLDER']) / filename.replace('.json', '.html')

    if not html_path.exists():
        return "Relatório não encontrado", 404

    with open(html_path, 'r', encoding='utf-8') as f:
        return f.read()

# APIs de Gabarito
@app.route('/api/gabaritos', methods=['GET'])
def list_gabaritos():
    """Lista todos os gabaritos disponíveis"""
    gabaritos = []
    gabaritos_path = Path(app.config['GABARITOS_FOLDER'])

    for json_file in sorted(gabaritos_path.glob('*.json')):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            is_oficial = json_file.name == GABARITO_OFICIAL
            gabaritos.append({
                'nome': json_file.stem,
                'arquivo': json_file.name,
                'questoes': len(data),
                'oficial': is_oficial
            })
        except:
            pass

    return jsonify(gabaritos)

@app.route('/api/gabarito/atual', methods=['GET'])
def get_gabarito_atual():
    """Retorna o gabarito oficial atual"""
    gabarito_path = Path(GABARITO_OFICIAL)

    if not gabarito_path.exists():
        return jsonify({'error': 'Nenhum gabarito selecionado'}), 404

    try:
        with open(gabarito_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except:
        return jsonify({'error': 'Erro ao ler gabarito'}), 500

@app.route('/api/gabarito/criar', methods=['POST'])
def criar_gabarito():
    """Cria um novo gabarito"""
    data = request.get_json()

    if not data:
        return jsonify({'error': 'Dados inválidos'}), 400

    nome = data.get('nome', '').strip()
    questoes = data.get('questoes', [])

    if not nome:
        return jsonify({'error': 'Nome do gabarito é obrigatório'}), 400

    if not questoes or len(questoes) == 0:
        return jsonify({'error': 'Adicione pelo menos uma questão'}), 400

    # Validar questões
    gabarito_dict = {}
    try:
        for i, resp in enumerate(questoes, 1):
            if resp.upper() not in ['A', 'B', 'C', 'D', 'E']:
                return jsonify({'error': f'Resposta inválida na questão {i}'}), 400
            gabarito_dict[i] = resp.upper()
    except:
        return jsonify({'error': 'Formato inválido nas questões'}), 400

    # Salvar gabarito
    gabarito_path = Path(app.config['GABARITOS_FOLDER']) / f"{nome}.json"

    try:
        with open(gabarito_path, 'w', encoding='utf-8') as f:
            json.dump(gabarito_dict, f, indent=2)
        return jsonify({'success': True, 'message': f'Gabarito "{nome}" criado com {len(gabarito_dict)} questões'})
    except Exception as e:
        return jsonify({'error': f'Erro ao salvar gabarito: {str(e)}'}), 500

@app.route('/api/gabarito/selecionar/<nome>', methods=['POST'])
def selecionar_gabarito(nome):
    """Seleciona um gabarito como oficial"""
    gabarito_path = Path(app.config['GABARITOS_FOLDER']) / f"{nome}.json"

    if not gabarito_path.exists():
        return jsonify({'error': 'Gabarito não encontrado'}), 404

    try:
        # Copiar para gabarito_oficial.json
        with open(gabarito_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        with open(GABARITO_OFICIAL, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)

        return jsonify({'success': True, 'message': f'Gabarito "{nome}" selecionado como oficial'})
    except Exception as e:
        return jsonify({'error': f'Erro ao selecionar gabarito: {str(e)}'}), 500

@app.route('/api/gabarito/deletar/<nome>', methods=['DELETE'])
def deletar_gabarito(nome):
    """Deleta um gabarito"""
    gabarito_path = Path(app.config['GABARITOS_FOLDER']) / f"{nome}.json"

    if not gabarito_path.exists():
        return jsonify({'error': 'Gabarito não encontrado'}), 404

    try:
        gabarito_path.unlink()
        return jsonify({'success': True, 'message': f'Gabarito "{nome}" deletado'})
    except Exception as e:
        return jsonify({'error': f'Erro ao deletar gabarito: {str(e)}'}), 500

@app.route('/api/gabaritos/limpar', methods=['DELETE'])
def limpar_gabaritos():
    """Deleta todos os gabaritos JSON"""
    gabaritos_path = Path(app.config['GABARITOS_FOLDER'])

    try:
        count = 0
        # Deletar todos os arquivos JSON na pasta de gabaritos
        for json_file in gabaritos_path.glob('*.json'):
            json_file.unlink()
            count += 1

        # Deletar também o gabarito oficial se existir
        gabarito_oficial_path = Path(GABARITO_OFICIAL)
        if gabarito_oficial_path.exists():
            gabarito_oficial_path.unlink()
            count += 1

        if count == 0:
            return jsonify({'success': True, 'message': 'Nenhum gabarito para deletar'})

        return jsonify({'success': True, 'message': f'{count} gabarito(s) deletado(s) com sucesso'})
    except Exception as e:
        return jsonify({'error': f'Erro ao limpar gabaritos: {str(e)}'}), 500

@app.route('/api/gabarito/gerar-pdf', methods=['POST'])
def gerar_gabarito_pdf():
    """Gera um gabarito em PDF"""
    data = request.get_json()

    if not data:
        return jsonify({'error': 'Dados inválidos'}), 400

    # Extrair parâmetros
    nome_arquivo = data.get('nome_arquivo', '').strip()
    titulo = data.get('titulo', 'GABARITO DE PROVA').strip()
    disciplina = data.get('disciplina', '').strip() or None
    professor = data.get('professor', '').strip() or None
    codigo_prova = data.get('codigo_prova', '').strip() or None
    num_questoes = int(data.get('num_questoes', 40))
    alternativas_str = data.get('alternativas', 'A,B,C,D,E').strip()

    # Validações
    if not nome_arquivo:
        return jsonify({'error': 'Nome do arquivo é obrigatório'}), 400

    if num_questoes < 1 or num_questoes > 200:
        return jsonify({'error': 'Número de questões deve estar entre 1 e 200'}), 400

    # Processar alternativas
    alternativas = [a.strip().upper() for a in alternativas_str.split(',') if a.strip()]
    if not alternativas:
        alternativas = ['A', 'B', 'C', 'D', 'E']

    # Adicionar .pdf se não tiver
    if not nome_arquivo.endswith('.pdf'):
        nome_arquivo += '.pdf'

    # Caminho completo
    pdf_path = Path(app.config['GABARITOS_PDF_FOLDER']) / nome_arquivo

    try:
        # Gerar PDF
        gerador = GeradorGabarito(str(pdf_path))
        gerador.gerar_gabarito_padrao(
            num_questoes=num_questoes,
            alternativas=alternativas,
            titulo=titulo,
            disciplina=disciplina,
            professor=professor,
            codigo_prova=codigo_prova
        )

        return jsonify({
            'success': True,
            'message': f'Gabarito PDF "{nome_arquivo}" gerado com sucesso!',
            'arquivo': nome_arquivo,
            'caminho': str(pdf_path)
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao gerar PDF: {str(e)}'}), 500

@app.route('/api/gabarito/download-pdf/<path:filename>')
def download_gabarito_pdf(filename):
    """Faz download de um gabarito PDF gerado"""
    pdf_path = Path(app.config['GABARITOS_PDF_FOLDER']) / filename

    if not pdf_path.exists():
        return jsonify({'error': 'PDF não encontrado'}), 404

    # Nome do arquivo para download (sem o caminho)
    download_name = Path(filename).name

    return send_file(pdf_path, as_attachment=True, download_name=download_name)

@app.route('/api/gabaritos-pdf', methods=['GET'])
def list_gabaritos_pdf():
    """Lista todos os gabaritos PDF gerados (incluindo subpastas)"""
    gabaritos_pdf = []
    gabaritos_pdf_path = Path(app.config['GABARITOS_PDF_FOLDER'])

    # PDFs diretos na pasta principal
    for pdf_file in gabaritos_pdf_path.glob('*.pdf'):
        stat = pdf_file.stat()
        gabaritos_pdf.append({
            'nome': pdf_file.name,
            'caminho_relativo': pdf_file.name,
            'tamanho': round(stat.st_size / 1024, 2),  # KB
            'data_criacao': datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M')
        })

    # PDFs em subpastas (gabaritos personalizados)
    for pdf_file in gabaritos_pdf_path.glob('*/*.pdf'):
        stat = pdf_file.stat()
        caminho_relativo = f"{pdf_file.parent.name}/{pdf_file.name}"
        gabaritos_pdf.append({
            'nome': f"{pdf_file.parent.name}",  # Nome da turma
            'caminho_relativo': caminho_relativo,
            'tamanho': round(stat.st_size / 1024, 2),  # KB
            'data_criacao': datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M')
        })

    # Ordenar por data (mais recentes primeiro)
    gabaritos_pdf.sort(key=lambda x: x['data_criacao'], reverse=True)

    return jsonify(gabaritos_pdf)

@app.route('/api/gabaritos-pdf/limpar', methods=['DELETE'])
def limpar_gabaritos_pdf():
    """Deleta todos os PDFs de gabaritos gerados"""
    gabaritos_pdf_path = Path(app.config['GABARITOS_PDF_FOLDER'])

    try:
        count = 0

        # Deletar PDFs diretos na pasta principal
        for pdf_file in gabaritos_pdf_path.glob('*.pdf'):
            pdf_file.unlink()
            count += 1

        # Deletar subpastas inteiras (gabaritos personalizados)
        for subpasta in gabaritos_pdf_path.iterdir():
            if subpasta.is_dir():
                # Deletar todos os arquivos dentro da subpasta
                for arquivo in subpasta.rglob('*'):
                    if arquivo.is_file():
                        arquivo.unlink()
                        count += 1

                # Deletar a subpasta vazia
                subpasta.rmdir()

        if count == 0:
            return jsonify({'success': True, 'message': 'Nenhum PDF para deletar'})

        return jsonify({'success': True, 'message': f'{count} arquivo(s) deletado(s) com sucesso'})
    except Exception as e:
        return jsonify({'error': f'Erro ao limpar PDFs: {str(e)}'}), 500

@app.route('/api/csv/upload', methods=['POST'])
def upload_csv():
    """Faz upload de CSV e analisa turmas"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Arquivo vazio'}), 400

    if not allowed_csv_file(file.filename):
        return jsonify({'error': 'Apenas arquivos CSV são permitidos'}), 400

    # Salvar arquivo
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['CSV_UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        # Ler alunos do CSV
        alunos = ler_csv_alunos(filepath)

        # Organizar por turmas
        turmas = listar_turmas(alunos)

        # Preparar resposta
        turmas_info = []
        for turma, alunos_turma in sorted(turmas.items()):
            turmas_info.append({
                'nome': turma,
                'num_alunos': len(alunos_turma)
            })

        return jsonify({
            'success': True,
            'arquivo': filename,
            'total_alunos': len(alunos),
            'total_turmas': len(turmas),
            'turmas': turmas_info
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao processar CSV: {str(e)}'}), 500

@app.route('/api/csv/gerar-gabaritos', methods=['POST'])
def gerar_gabaritos_csv():
    """Gera gabaritos personalizados a partir do CSV"""
    data = request.get_json()

    if not data:
        return jsonify({'error': 'Dados inválidos'}), 400

    # Extrair parâmetros
    csv_filename = data.get('csv_filename', '').strip()
    turmas_selecionadas = data.get('turmas', [])
    titulo = data.get('titulo', 'GABARITO DE PROVA').strip()
    disciplina = data.get('disciplina', '').strip() or None
    professor = data.get('professor', '').strip() or None
    codigo_prova = data.get('codigo_prova', '').strip() or None
    num_questoes = int(data.get('num_questoes', 40))
    alternativas_str = data.get('alternativas', 'A,B,C,D,E').strip()

    # Validações
    if not csv_filename:
        return jsonify({'error': 'Nome do arquivo CSV é obrigatório'}), 400

    if not turmas_selecionadas or len(turmas_selecionadas) == 0:
        return jsonify({'error': 'Selecione pelo menos uma turma'}), 400

    csv_path = Path(app.config['CSV_UPLOAD_FOLDER']) / csv_filename
    if not csv_path.exists():
        return jsonify({'error': 'Arquivo CSV não encontrado'}), 404

    # Processar alternativas
    alternativas = [a.strip().upper() for a in alternativas_str.split(',') if a.strip()]
    if not alternativas:
        alternativas = ['A', 'B', 'C', 'D', 'E']

    try:
        # Ler alunos do CSV
        alunos = ler_csv_alunos(str(csv_path))

        # Organizar por turmas
        turmas = listar_turmas(alunos)

        # Configuração
        config = {
            'titulo': titulo,
            'disciplina': disciplina,
            'professor': professor,
            'codigo_prova': codigo_prova,
            'num_questoes': num_questoes,
            'alternativas': alternativas
        }

        # Pasta de saída
        pasta_saida = app.config['GABARITOS_PDF_FOLDER']

        # Gerar gabaritos para as turmas selecionadas
        total_alunos = 0
        pdfs_gerados = []

        for turma_nome in turmas_selecionadas:
            if turma_nome in turmas:
                alunos_turma = turmas[turma_nome]
                total_alunos += len(alunos_turma)

                # Gerar gabarito para a turma
                gerar_gabaritos_turma(alunos_turma, turma_nome, pasta_saida, config)

                # Nome do PDF gerado
                import re
                nome_pasta_limpo = re.sub(r'[/\\:*?"<>|]', '', turma_nome).strip()
                nome_pasta_limpo = re.sub(r'\s+', '_', nome_pasta_limpo)
                pdf_nome = f"{nome_pasta_limpo}/{nome_pasta_limpo}.pdf"
                pdfs_gerados.append({
                    'turma': turma_nome,
                    'arquivo': pdf_nome,
                    'num_alunos': len(alunos_turma)
                })

        return jsonify({
            'success': True,
            'message': f'Gabaritos gerados com sucesso para {len(turmas_selecionadas)} turma(s)!',
            'total_alunos': total_alunos,
            'total_turmas': len(turmas_selecionadas),
            'pdfs': pdfs_gerados
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao gerar gabaritos: {str(e)}'}), 500

# Funções auxiliares para mapeamento de alunos
def carregar_mapeamento_alunos():
    """Carrega mapeamento de Nome → Turma do CSV de referência"""
    csv_ref_path = Path(app.config['CSV_ALUNOS_FOLDER']) / 'alunos_referencia.csv'

    if not csv_ref_path.exists():
        return {}

    mapeamento = {}
    try:
        with open(csv_ref_path, 'r', encoding='utf-8') as f:
            # Ler CSV
            linhas = f.readlines()

            # Encontrar linha de cabeçalho (ignorar linhas vazias e cabeçalho de data)
            idx_cabecalho = -1
            for i, linha in enumerate(linhas):
                if 'Aluno' in linha and 'Turma' in linha:
                    idx_cabecalho = i
                    break

            if idx_cabecalho == -1:
                return {}

            # Processar dados
            cabecalho = linhas[idx_cabecalho].strip().split(';')
            idx_aluno = cabecalho.index('Aluno') if 'Aluno' in cabecalho else 0
            idx_turma = cabecalho.index('Turma') if 'Turma' in cabecalho else 8

            for linha in linhas[idx_cabecalho + 1:]:
                linha = linha.strip()
                if not linha:
                    continue

                partes = linha.split(';')
                if len(partes) > max(idx_aluno, idx_turma):
                    nome = partes[idx_aluno].strip().upper()
                    turma = partes[idx_turma].strip()
                    mapeamento[nome] = turma
    except Exception as e:
        print(f"Erro ao carregar mapeamento de alunos: {e}")

    return mapeamento

def encontrar_turma_aluno(nome_pdf):
    """Tenta encontrar a turma do aluno baseado no nome do PDF"""
    mapeamento = carregar_mapeamento_alunos()

    if not mapeamento:
        return None

    # Normalizar nome do PDF (remover extensão, underscores, etc)
    nome_normalizado = nome_pdf.replace('.pdf', '').replace('_', ' ').strip().upper()

    # Tentar match exato
    if nome_normalizado in mapeamento:
        return mapeamento[nome_normalizado]

    # Tentar match parcial
    for nome_aluno, turma in mapeamento.items():
        if nome_aluno in nome_normalizado or nome_normalizado in nome_aluno:
            return turma

    return None

@app.route('/api/csv-alunos/upload', methods=['POST'])
def upload_csv_alunos():
    """Faz upload do CSV de referência de alunos"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Arquivo vazio'}), 400

    if not allowed_csv_file(file.filename):
        return jsonify({'error': 'Apenas arquivos CSV são permitidos'}), 400

    # Salvar arquivo como alunos_referencia.csv
    filepath = Path(app.config['CSV_ALUNOS_FOLDER']) / 'alunos_referencia.csv'
    file.save(filepath)

    try:
        # Carregar e analisar
        mapeamento = carregar_mapeamento_alunos()
        turmas_unicas = sorted(set(mapeamento.values()))

        return jsonify({
            'success': True,
            'message': f'CSV carregado com sucesso!',
            'total_alunos': len(mapeamento),
            'total_turmas': len(turmas_unicas),
            'turmas': turmas_unicas
        })

    except Exception as e:
        return jsonify({'error': f'Erro ao processar CSV: {str(e)}'}), 500

@app.route('/api/csv-alunos/status', methods=['GET'])
def status_csv_alunos():
    """Retorna status do CSV de referência"""
    csv_ref_path = Path(app.config['CSV_ALUNOS_FOLDER']) / 'alunos_referencia.csv'

    if not csv_ref_path.exists():
        return jsonify({'carregado': False})

    try:
        mapeamento = carregar_mapeamento_alunos()
        turmas_unicas = sorted(set(mapeamento.values()))

        return jsonify({
            'carregado': True,
            'total_alunos': len(mapeamento),
            'total_turmas': len(turmas_unicas),
            'turmas': turmas_unicas
        })
    except:
        return jsonify({'carregado': False})

@app.route('/api/csv-alunos/limpar', methods=['DELETE'])
def limpar_csv_alunos():
    """Remove o CSV de referência"""
    csv_ref_path = Path(app.config['CSV_ALUNOS_FOLDER']) / 'alunos_referencia.csv'

    try:
        if csv_ref_path.exists():
            csv_ref_path.unlink()
            return jsonify({'success': True, 'message': 'CSV de referência removido'})
        else:
            return jsonify({'success': True, 'message': 'Nenhum CSV para remover'})
    except Exception as e:
        return jsonify({'error': f'Erro ao remover CSV: {str(e)}'}), 500

@app.route('/api/gabaritos-pdf/download-todos', methods=['GET'])
def download_todos_pdfs():
    """Baixa todos os PDFs de gabaritos em um arquivo ZIP"""
    gabaritos_pdf_path = Path(app.config['GABARITOS_PDF_FOLDER'])

    try:
        # Criar arquivo ZIP em memória
        memory_file = io.BytesIO()

        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Contar arquivos
            count = 0

            # Adicionar PDFs diretos na pasta principal
            for pdf_file in gabaritos_pdf_path.glob('*.pdf'):
                zf.write(pdf_file, pdf_file.name)
                count += 1

            # Adicionar PDFs em subpastas (gabaritos personalizados)
            for pdf_file in gabaritos_pdf_path.glob('*/*.pdf'):
                # Manter estrutura de pastas no ZIP
                arcname = f"{pdf_file.parent.name}/{pdf_file.name}"
                zf.write(pdf_file, arcname)
                count += 1

        if count == 0:
            return jsonify({'error': 'Nenhum PDF encontrado para baixar'}), 404

        # Retornar para o início do arquivo
        memory_file.seek(0)

        # Nome do arquivo ZIP com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_filename = f'gabaritos_{timestamp}.zip'

        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )

    except Exception as e:
        return jsonify({'error': f'Erro ao criar ZIP: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False)
